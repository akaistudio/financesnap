"""
Varnam Suite – Central Hub FinanceSnap – The Varnam Suite Hub Mini-ERP
Central dashboard & mini-ERP for 1-5 person companies.
Auto-receives company registrations from all Varnam Suite apps.
"""
import os, hashlib, json, requests, secrets
import bcrypt, smtplib
import hmac, base64, time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from functools import wraps
from urllib.parse import quote as urlquote
from flask import Flask, request, jsonify, redirect, url_for, session, render_template, flash, send_file
import psycopg2, psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'snapsuite-hub-2026')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=90)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# ── SSO ─────────────────────────────────────────────────────────
VARNAM_SSO_SECRET = os.environ.get('VARNAM_SSO_SECRET', 'varnam-suite-sso-2026')

def generate_sso_token(email):
    """Generate a short-lived signed token for cross-app login."""
    timestamp = str(int(time.time()))
    payload = f"{email}|{timestamp}"
    sig = hmac.new(VARNAM_SSO_SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()
    token = base64.urlsafe_b64encode(f"{payload}|{sig}".encode()).decode()
    return token

MAX_COMPANIES = 500

def get_db():
    conn = psycopg2.connect(os.environ['DATABASE_URL'], cursor_factory=psycopg2.extras.RealDictCursor)
    conn.autocommit = True
    return conn

def init_db():
    conn = get_db(); cur = conn.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS users (
        id SERIAL PRIMARY KEY, email TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL,
        name TEXT DEFAULT '', currency TEXT DEFAULT 'INR',
        is_superadmin BOOLEAN DEFAULT FALSE, created_at TIMESTAMP DEFAULT NOW()
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS companies (
        id SERIAL PRIMARY KEY, name TEXT NOT NULL, currency TEXT DEFAULT 'INR',
        owner_email TEXT DEFAULT '', created_at TIMESTAMP DEFAULT NOW()
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS company_apps (
        id SERIAL PRIMARY KEY,
        company_id INTEGER REFERENCES companies(id) ON DELETE CASCADE,
        app_name TEXT NOT NULL, app_company_name TEXT DEFAULT '',
        app_url TEXT DEFAULT '', registered_at TIMESTAMP DEFAULT NOW(),
        UNIQUE(company_id, app_name)
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS company_users (
        id SERIAL PRIMARY KEY,
        company_id INTEGER REFERENCES companies(id) ON DELETE CASCADE,
        user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
        role TEXT DEFAULT 'owner', UNIQUE(company_id, user_id)
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS app_settings (
        key TEXT PRIMARY KEY, value TEXT DEFAULT ''
    )''')
    for m in [
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS is_superadmin BOOLEAN DEFAULT FALSE",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS name TEXT DEFAULT ''",
        "ALTER TABLE companies ADD COLUMN IF NOT EXISTS owner_email TEXT DEFAULT ''",
        "UPDATE users SET is_superadmin = TRUE WHERE id = (SELECT MIN(id) FROM users)",
        """CREATE TABLE IF NOT EXISTS otp_codes (
            id SERIAL PRIMARY KEY, email TEXT NOT NULL, code TEXT NOT NULL,
            purpose TEXT DEFAULT 'login', attempts INTEGER DEFAULT 0,
            used BOOLEAN DEFAULT FALSE, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP NOT NULL)""",
        "CREATE INDEX IF NOT EXISTS idx_otp_email ON otp_codes(email, purpose, used)",
        """CREATE TABLE IF NOT EXISTS bank_statements (
            id SERIAL PRIMARY KEY,
            user_id INTEGER REFERENCES users(id),
            company_name TEXT DEFAULT '',
            filename TEXT DEFAULT '',
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            row_count INTEGER DEFAULT 0,
            matched_count INTEGER DEFAULT 0,
            currency TEXT DEFAULT 'USD'
        )""",
        """CREATE TABLE IF NOT EXISTS bank_transactions (
            id SERIAL PRIMARY KEY,
            statement_id INTEGER REFERENCES bank_statements(id) ON DELETE CASCADE,
            txn_date TEXT DEFAULT '',
            description TEXT DEFAULT '',
            amount DOUBLE PRECISION DEFAULT 0,
            txn_type TEXT DEFAULT 'debit',
            status TEXT DEFAULT 'unmatched',
            matched_type TEXT DEFAULT '',
            matched_id TEXT DEFAULT '',
            expense_snap_id TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
    ]:
        try: cur.execute(m)
        except: pass
    conn.close()

init_db()

# ── Helpers ─────────────────────────────────────────────────────
def hash_pw(p):
    return bcrypt.hashpw(p.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def check_pw(pw, hashed):
    try:
        return bcrypt.checkpw(pw.encode('utf-8'), hashed.encode('utf-8'))
    except (ValueError, AttributeError):
        if hashlib.sha256(pw.encode()).hexdigest() == hashed:
            return True
        return False

def generate_otp():
    return f"{secrets.randbelow(900000) + 100000}"

def send_otp_email(email, code, purpose='login'):
    resend_key = os.environ.get('RESEND_API_KEY', '')
    from_email = os.environ.get('SMTP_FROM', 'noreply@usevarnam.com')
    purpose_text = 'login' if purpose == 'login' else 'verification'
    subject = f"Your FinanceSnap {purpose_text} code: {code}"
    html = f"""<div style="font-family:sans-serif;max-width:400px;margin:0 auto;padding:24px">
        <h2 style="color:#10b981">FinanceSnap</h2>
        <p style="color:#666;font-size:14px">Your {purpose_text} code is:</p>
        <div style="font-size:36px;font-weight:800;letter-spacing:8px;color:#1a1a2e;text-align:center;
                    padding:20px;background:#f0f4ff;border-radius:12px;margin:16px 0">{code}</div>
        <p style="color:#999;font-size:12px">This code expires in 5 minutes. Do not share it.</p>
        <p style="color:#999;font-size:11px;margin-top:20px">Part of <a href="https://snapsuite.up.railway.app" style="color:#10b981">Varnam Suite</a></p>
    </div>"""
    if not resend_key:
        print(f"⚠️ RESEND_API_KEY not set. OTP for {email}: {code}")
        return False
    import requests as http_req
    try:
        r = http_req.post('https://api.resend.com/emails', json={
            'from': from_email, 'to': [email], 'subject': subject, 'html': html
        }, headers={'Authorization': f'Bearer {resend_key}'}, timeout=10)
        if r.status_code == 200:
            print(f"✅ OTP sent to {email}")
            return True
        else:
            print(f"❌ Resend error {r.status_code}: {r.text}")
            print(f"💡 OTP for {email}: {code}")
            return False
    except Exception as e:
        print(f"❌ Email failed: {e}")
        print(f"💡 OTP for {email}: {code}")
        return False


def cs(c): return {'INR':'₹','USD':'$','EUR':'€','GBP':'£','CAD':'C$','MYR':'RM','SGD':'S$','AUD':'A$','AED':'AED'}.get(c, c+' ')

def login_required(f):
    @wraps(f)
    def decorated(*a, **kw):
        if 'user_id' not in session: return redirect(url_for('login'))
        return f(*a, **kw)
    return decorated

def get_user():
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT * FROM users WHERE id=%s', (session['user_id'],))
    u = cur.fetchone(); conn.close(); return u

def get_user_companies(user):
    conn = get_db(); cur = conn.cursor()
    if user['is_superadmin']:
        cur.execute('SELECT * FROM companies ORDER BY name')
    else:
        cur.execute('''SELECT c.* FROM companies c
                      JOIN company_users cu ON c.id=cu.company_id
                      WHERE cu.user_id=%s ORDER BY c.name''', (user['id'],))
    r = cur.fetchall(); conn.close(); return r

def get_company_apps(cid):
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT * FROM company_apps WHERE company_id=%s', (cid,))
    r = cur.fetchall(); conn.close()
    apps = {a['app_name']: a for a in r}
    # ProposalSnap is always available (standalone tool, no per-company registration)
    if 'ProposalSnap' not in apps:
        apps['ProposalSnap'] = {'app_name': 'ProposalSnap', 'app_url': APP_URLS.get('ProposalSnap', '')}
    return apps

DEFAULT_URLS = {
    'ExpenseSnap': os.environ.get('EXPENSESNAP_URL', 'https://expensesnap.up.railway.app'),
    'InvoiceSnap': os.environ.get('INVOICESNAP_URL', 'https://invoicesnap.up.railway.app'),
    'ContractSnap': os.environ.get('CONTRACTSNAP_URL', 'https://contractsnap-app.up.railway.app'),
    'PayslipSnap': os.environ.get('PAYSLIPSNAP_URL', 'https://payslipsnap.up.railway.app'),
    'ProposalSnap': os.environ.get('PROPOSALSNAP_URL', 'https://proposalsnap.up.railway.app'),
    'SplitSnap': os.environ.get('SPLITSNAP_URL', 'https://splitsnap.up.railway.app'),
}

def get_app_urls():
    urls = dict(DEFAULT_URLS)
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT key, value FROM app_settings WHERE key LIKE 'url_%'")
        for row in cur.fetchall():
            app_name = row['key'].replace('url_', '')
            if row['value'].strip():
                urls[app_name] = row['value'].strip()
        conn.close()
    except: pass
    return urls

# For backward compat
APP_URLS = DEFAULT_URLS

@app.before_request
def load_app_urls():
    global APP_URLS
    APP_URLS = get_app_urls()

def fetch_api(base_url, endpoint, api_key):
    if not base_url: return None
    try:
        r = requests.get(base_url.rstrip('/') + endpoint,
                        headers={'X-API-Key': api_key}, timeout=30)
        if r.status_code == 200: return r.json()
    except: pass
    return None

@app.route('/api/test-connections')
def test_connections():
    """Public endpoint to debug API connections."""
    results = {}
    # Try common emails
    test_emails = set()
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute('SELECT email FROM users ORDER BY id LIMIT 5')
        for u in cur.fetchall(): test_emails.add(u['email'])
        conn.close()
    except: pass

    if not test_emails:
        test_emails = {'test@example.com'}

    for email in test_emails:
        results[email] = {}
        tests = {
            'ExpenseSnap /api/companies/external': (APP_URLS['ExpenseSnap'], '/api/companies/external'),
            'ExpenseSnap /api/expenses/external': (APP_URLS['ExpenseSnap'], '/api/expenses/external'),
            'InvoiceSnap /api/invoices': (APP_URLS['InvoiceSnap'], '/api/invoices'),
            'ContractSnap /api/contracts': (APP_URLS['ContractSnap'], '/api/contracts'),
            'PayslipSnap /api/payroll': (APP_URLS['PayslipSnap'], '/api/payroll'),
        }
        for label, (base, ep) in tests.items():
            try:
                r = requests.get(base.rstrip('/') + ep,
                    headers={'X-API-Key': email}, timeout=30)
                results[email][label] = {
                    'status': r.status_code,
                    'body': r.text[:200]
                }
            except Exception as e:
                results[email][label] = {'status': 'error', 'body': str(e)[:200]}

    return jsonify({
        'app_urls': APP_URLS,
        'test_emails': list(test_emails),
        'results': results
    })

# ── Central API ─────────────────────────────────────────────────
@app.route('/api/register-company', methods=['POST'])
def api_register_company():
    """Called by any Varnam Suite app when a user creates a company.
    {app_name, company_name, email, currency, app_url}"""
    data = request.json or {}
    app_name = data.get('app_name', '').strip()
    company_name = data.get('company_name', '').strip()
    email = data.get('email', '').strip().lower()
    currency = data.get('currency', 'INR').upper()
    app_url = data.get('app_url', '').strip()

    if not company_name or not app_name:
        return jsonify({'error': 'company_name and app_name required'}), 400

    conn = get_db(); cur = conn.cursor()

    # Company limit
    cur.execute('SELECT COUNT(*) as cnt FROM companies')
    if cur.fetchone()['cnt'] >= MAX_COMPANIES:
        conn.close()
        return jsonify({'error': f'Max {MAX_COMPANIES} companies reached'}), 400

    # Find or create company
    cur.execute('SELECT * FROM companies WHERE LOWER(name)=LOWER(%s)', (company_name,))
    company = cur.fetchone()
    if not company:
        cur.execute('INSERT INTO companies (name,currency,owner_email) VALUES (%s,%s,%s) RETURNING *',
                   (company_name, currency, email))
        company = cur.fetchone()

    # Register app link
    cur.execute('''INSERT INTO company_apps (company_id,app_name,app_company_name,app_url)
                  VALUES (%s,%s,%s,%s) ON CONFLICT (company_id,app_name)
                  DO UPDATE SET app_company_name=EXCLUDED.app_company_name, app_url=EXCLUDED.app_url''',
               (company['id'], app_name, company_name, app_url))

    # Link user if exists
    if email:
        cur.execute('SELECT id FROM users WHERE email=%s', (email,))
        u = cur.fetchone()
        if u:
            cur.execute('''INSERT INTO company_users (company_id,user_id,role)
                          VALUES (%s,%s,'owner') ON CONFLICT DO NOTHING''', (company['id'], u['id']))
    conn.close()
    return jsonify({'success': True, 'company_id': company['id'], 'company_name': company['name']})

@app.route('/api/companies')
def api_list_companies():
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT id,name,currency FROM companies ORDER BY name')
    r = cur.fetchall(); conn.close()
    return jsonify({'companies': r})

# ── Auth ────────────────────────────────────────────────────────
@app.route('/demo')
def demo_login():
    """One-click demo login with pre-loaded Bloom Studio data."""
    demo_email = 'demo@snapsuite.app'
    demo_pw = hash_pw('demo123')
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT * FROM users WHERE email=%s', (demo_email,))
    user = cur.fetchone()
    if not user:
        cur.execute('''INSERT INTO users (email,password_hash,name,currency,is_superadmin)
                      VALUES (%s,%s,%s,%s,FALSE) RETURNING *''',
                   (demo_email, demo_pw, 'Demo User', 'INR'))
        user = cur.fetchone()
    session['user_id'] = user['id']
    cur.execute("SELECT * FROM companies WHERE LOWER(name)='bloom studio' AND owner_email=%s", (demo_email,))
    company = cur.fetchone()
    if not company:
        cur.execute("INSERT INTO companies (name,currency,owner_email) VALUES ('Bloom Studio','INR',%s) RETURNING *", (demo_email,))
        company = cur.fetchone()
        for app_name in ['ExpenseSnap', 'InvoiceSnap', 'ContractSnap', 'PayslipSnap']:
            url = APP_URLS.get(app_name, '')
            cur.execute('''INSERT INTO company_apps (company_id,app_name,app_company_name,app_url)
                          VALUES (%s,%s,'Bloom Studio',%s) ON CONFLICT DO NOTHING''',
                       (company['id'], app_name, url))
    cur.execute('INSERT INTO company_users (company_id,user_id,role) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING',
               (company['id'], user['id'], 'owner'))
    conn.close()
    demo_secret = 'snapsuite-demo-2026'
    for app_name in ['ExpenseSnap', 'InvoiceSnap', 'ContractSnap', 'PayslipSnap']:
        url = APP_URLS.get(app_name, '')
        if url:
            try: requests.post(url.rstrip('/') + '/api/demo-setup', headers={'X-Demo-Secret': demo_secret}, timeout=30)
            except: pass
    return redirect(url_for('dashboard'))

@app.route('/welcome')
def welcome():
    if 'user_id' in session:
        return redirect('/')
    return render_template('login.html')

@app.route('/login', methods=['GET'])
def login():
    if 'user_id' in session: return redirect('/')
    return render_template('login.html')

@app.route('/register', methods=['GET'])
def register():
    if 'user_id' in session: return redirect('/')
    return render_template('login.html', show_register=True)

@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('login'))

def auto_link_companies(cur, email, user_id):
    """Auto-link user to companies by owner email"""
    cur.execute('SELECT id FROM companies WHERE owner_email=%s', (email,))
    for c in cur.fetchall():
        cur.execute('''INSERT INTO company_users (company_id,user_id,role)
                      VALUES (%s,%s,'owner') ON CONFLICT DO NOTHING''', (c['id'], user_id))

# --- OTP API ---
@app.route('/api/auth/send-otp', methods=['POST'])
def send_otp():
    data = request.get_json()
    email = (data.get('email') or '').strip().lower()
    purpose = data.get('purpose', 'login')
    if not email or '@' not in email:
        return jsonify({"error": "Valid email required"}), 400
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""SELECT COUNT(*) as cnt FROM otp_codes
                   WHERE email=%s AND created_at > NOW() - INTERVAL '15 minutes'""", (email,))
    if cur.fetchone()['cnt'] >= 5:
        conn.close()
        return jsonify({"error": "Too many requests. Wait 15 minutes."}), 429
    if purpose == 'login':
        cur.execute('SELECT id FROM users WHERE email=%s', (email,))
        if not cur.fetchone():
            conn.close()
            return jsonify({"error": "No account found with this email"}), 404
    if purpose == 'register':
        cur.execute('SELECT id FROM users WHERE email=%s', (email,))
        if cur.fetchone():
            conn.close()
            return jsonify({"error": "Email already registered. Please sign in."}), 409
    cur.execute("UPDATE otp_codes SET used=TRUE WHERE email=%s AND purpose=%s AND used=FALSE", (email, purpose))
    code = generate_otp()
    expires = datetime.utcnow() + timedelta(minutes=5)
    cur.execute("INSERT INTO otp_codes (email, code, purpose, expires_at) VALUES (%s,%s,%s,%s)",
                (email, code, purpose, expires))
    conn.close()
    if send_otp_email(email, code, purpose):
        return jsonify({"success": True})
    return jsonify({"success": True, "fallback_code": code, "email_failed": True})

@app.route('/api/auth/verify-otp', methods=['POST'])
def verify_otp():
    data = request.get_json()
    email = (data.get('email') or '').strip().lower()
    code = (data.get('code') or '').strip()
    purpose = data.get('purpose', 'login')
    if not email or not code or len(code) != 6:
        return jsonify({"error": "Email and 6-digit code required"}), 400
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""SELECT * FROM otp_codes
                   WHERE email=%s AND purpose=%s AND used=FALSE AND expires_at > NOW()
                   ORDER BY created_at DESC LIMIT 1""", (email, purpose))
    otp_rec = cur.fetchone()
    if not otp_rec:
        conn.close()
        return jsonify({"error": "Code expired. Request a new one."}), 400
    if otp_rec['attempts'] >= 3:
        cur.execute("UPDATE otp_codes SET used=TRUE WHERE id=%s", (otp_rec['id'],))
        conn.close()
        return jsonify({"error": "Too many attempts. Request a new code."}), 429
    cur.execute("UPDATE otp_codes SET attempts=attempts+1 WHERE id=%s", (otp_rec['id'],))
    if not secrets.compare_digest(code, otp_rec['code']):
        conn.close()
        remaining = 2 - otp_rec['attempts']
        return jsonify({"error": f"Invalid code. {remaining} attempt(s) remaining."}), 400
    cur.execute("UPDATE otp_codes SET used=TRUE WHERE id=%s", (otp_rec['id'],))
    if purpose == 'login':
        cur.execute('SELECT * FROM users WHERE email=%s', (email,))
        user = cur.fetchone()
        if user:
            session['user_id'] = user['id']
            session.permanent = True
            auto_link_companies(cur, email, user['id'])
            conn.close()
            return jsonify({"success": True, "redirect": "/"})
        conn.close()
        return jsonify({"error": "User not found"}), 404
    conn.close()
    return jsonify({"success": True, "verified": True})

@app.route('/api/auth/register', methods=['POST'])
def api_register():
    data = request.get_json()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password', '') or secrets.token_hex(16)
    company = (data.get('company_name') or '').strip()
    currency = data.get('currency', 'INR')
    code = (data.get('code') or '').strip()
    if not email or not company:
        return jsonify({"error": "Email and company name required"}), 400
    if len(code) != 6:
        return jsonify({"error": "Valid 6-digit code required"}), 400
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""SELECT * FROM otp_codes
                   WHERE email=%s AND purpose='register' AND used=FALSE AND expires_at > NOW()
                   ORDER BY created_at DESC LIMIT 1""", (email,))
    otp_rec = cur.fetchone()
    if not otp_rec or not secrets.compare_digest(code, otp_rec['code']):
        conn.close()
        return jsonify({"error": "Invalid or expired code"}), 400
    if otp_rec['attempts'] >= 3:
        conn.close()
        return jsonify({"error": "Too many attempts. Request a new code."}), 429
    cur.execute("UPDATE otp_codes SET used=TRUE WHERE id=%s", (otp_rec['id'],))
    cur.execute('SELECT id FROM users WHERE email=%s', (email,))
    if cur.fetchone():
        conn.close()
        return jsonify({"error": "Email already registered"}), 409
    cur.execute('SELECT COUNT(*) as cnt FROM users')
    is_first = cur.fetchone()['cnt'] == 0
    try:
        cur.execute('''INSERT INTO users (email,password_hash,name,currency,is_superadmin)
                      VALUES (%s,%s,%s,%s,%s) RETURNING id''',
                   (email, hash_pw(password), company, currency, is_first))
        uid = cur.fetchone()['id']
        session['user_id'] = uid
        session.permanent = True
        auto_link_companies(cur, email, uid)
        conn.close()
        return jsonify({"success": True, "redirect": "/"})
    except psycopg2.IntegrityError:
        conn.close()
        return jsonify({"error": "Email already registered"}), 409

# ── App Hub (always accessible) ─────────────────────────────────
@app.route('/apps')
@login_required
def apps_hub():
    user = get_user()
    companies = get_user_companies(user)
    is_demo = user['email'] == 'demo@snapsuite.app'
    if is_demo:
        # Demo user goes to /demo on each app, no SSO needed
        final_urls = {k: v + '/demo' for k, v in APP_URLS.items()}
    else:
        # Real user gets SSO token baked into each URL
        sso_token = generate_sso_token(user['email'])
        final_urls = {k: v + '/auto-login?token=' + sso_token for k, v in APP_URLS.items()}
    return render_template('apps.html', user=user, app_urls=final_urls, companies=companies)

# ── Dashboard ───────────────────────────────────────────────────
@app.route('/')
def dashboard():
    if 'user_id' not in session:
        return redirect('/welcome')
    user = get_user()
    companies = get_user_companies(user)
    if not companies:
        return redirect(url_for('apps_hub'))

    sel_id = request.args.get('company', '')
    selected = None
    for c in companies:
        if str(c['id']) == sel_id: selected = c; break
    if not selected: selected = companies[0]

    curr = cs(selected.get('currency', 'INR'))
    apps = get_company_apps(selected['id'])
    api_key = selected.get('owner_email', user['email'])

    invoices=[]; contracts=[]; expenses=[]; payslips=[]

    if 'InvoiceSnap' in apps:
        url = apps['InvoiceSnap']['app_url'] or APP_URLS['InvoiceSnap']
        r = fetch_api(url, f'/api/invoices?company_name={urlquote(selected["name"])}', api_key)
        if r: invoices = r.get('invoices', [])

    if 'ContractSnap' in apps:
        url = apps['ContractSnap']['app_url'] or APP_URLS['ContractSnap']
        r = fetch_api(url, f'/api/contracts?company_name={urlquote(selected["name"])}', api_key)
        if r: contracts = r.get('contracts', [])

    if 'ExpenseSnap' in apps:
        url = apps['ExpenseSnap']['app_url'] or APP_URLS['ExpenseSnap']
        # Find matching expense company ID
        ecid = ''
        r2 = fetch_api(url, '/api/companies/external', api_key)
        if r2:
            for ec in r2.get('companies', []):
                if ec['name'].lower().strip() == selected['name'].lower().strip():
                    ecid = str(ec['id']); break
        ep = '/api/expenses/external'
        if ecid: ep += f'?company_id={ecid}'
        r = fetch_api(url, ep, api_key)
        if r: expenses = r.get('expenses', [])

    if 'PayslipSnap' in apps:
        url = apps['PayslipSnap']['app_url'] or APP_URLS['PayslipSnap']
        r = fetch_api(url, f'/api/payroll?company_name={urlquote(selected["name"])}', api_key)
        if r: payslips = r.get('payslips', [])

    # Metrics
    total_invoiced = sum(float(i.get('total',0) or 0) for i in invoices)
    total_paid = sum(float(i.get('total',0) or 0) for i in invoices if i.get('status')=='paid')
    total_unpaid = total_invoiced - total_paid
    total_overdue = sum(float(i.get('total',0) or 0) for i in invoices if i.get('status')=='overdue')
    total_expenses = sum(float(e.get('total',0) or e.get('amount',0) or 0) for e in expenses)
    # Employer cost = gross + employer PF + employer ESI (CTC, not net pay)
    total_payroll = sum(
        float(p.get('gross_earnings',0) or 0) +
        float(p.get('pf_employer',0) or 0) +
        float(p.get('esi_employer',0) or 0)
        for p in payslips)
    total_payroll_gross = sum(float(p.get('gross_earnings',0) or 0) for p in payslips)
    total_payroll_net = sum(float(p.get('net_pay',0) or 0) for p in payslips)
    contract_value = sum(float(c.get('total_value',0) or 0) for c in contracts)
    active_contracts = [c for c in contracts if c.get('status') in ('active','signed')]
    revenue = total_paid
    costs = total_expenses + total_payroll
    profit = revenue - costs

    # Expense categories
    cats = {}
    for e in expenses:
        c = e.get('category','Other') or 'Other'
        cats[c] = cats.get(c,0) + float(e.get('total',0) or e.get('amount',0) or 0)
    expense_cats = sorted(cats.items(), key=lambda x:-x[1])[:8]

    # Monthly chart — use paid_at for income (actual cash received date)
    monthly = []
    for i in range(5,-1,-1):
        d = datetime.now() - timedelta(days=30*i)
        ym = d.strftime('%Y-%m')
        inc = sum(float(inv.get('total',0) or 0) for inv in invoices
                  if inv.get('status')=='paid' and (
                      (inv.get('paid_at') or inv.get('date','') or '')).startswith(ym))
        exp = sum(float(e.get('total',0) or e.get('amount',0) or 0) for e in expenses
                  if (e.get('date','') or e.get('receipt_date','') or '').startswith(ym))
        pay = sum(float(p.get('net_pay',0) or 0) for p in payslips
                  if str(p.get('month','')).zfill(2)==d.strftime('%m') and str(p.get('year',''))==d.strftime('%Y'))
        monthly.append({'label':d.strftime('%b'),'income':inc,'expense':exp,'payroll':pay})
    mcv = max([m['income']+m['expense']+m['payroll'] for m in monthly]+[1])

    # ── Cash Flow Statement ──────────────────────────────────────
    # Operating Activities — payroll treated as immediate cash out
    cf_receipts      = total_paid
    cf_exp_paid      = total_expenses
    cf_payroll_paid  = total_payroll_net
    cf_employer_contrib = total_payroll - total_payroll_gross  # PF/ESI/CPF/SDL
    cf_net_operating = cf_receipts - cf_exp_paid - cf_payroll_paid - cf_employer_contrib

    cf_net_investing  = 0.0
    cf_net_financing  = 0.0
    cf_net_total      = cf_net_operating

    cash_flow = {
        'receipts':          cf_receipts,
        'exp_paid':          cf_exp_paid,
        'payroll_paid':      cf_payroll_paid,
        'employer_contrib':  cf_employer_contrib,
        'net_operating':     cf_net_operating,
        'net_investing':     cf_net_investing,
        'net_financing':     cf_net_financing,
        'net_total':         cf_net_total,
    }

    # ── Balance Sheet ────────────────────────────────────────────
    # ASSETS
    bs_ar           = total_unpaid
    bs_cash         = max(0.0, cf_net_operating)  # Cash proxy = what's left after all outflows
    bs_total_assets = bs_ar + bs_cash

    # LIABILITIES — payroll already paid, so no salary payable
    bs_total_liabilities = 0.0

    # EQUITY
    bs_retained_earnings = profit
    bs_total_equity      = bs_total_assets - bs_total_liabilities

    balance_sheet = {
        'cash':              bs_cash,
        'ar':                bs_ar,
        'total_assets':      bs_total_assets,
        'total_liabilities': bs_total_liabilities,
        'retained_earnings': bs_retained_earnings,
        'total_equity':      bs_total_equity,
    }

    is_demo = user['email'] == 'demo@snapsuite.app'

    # Proposals — pull from ProposalSnap API
    proposals = []
    if 'ProposalSnap' in apps:
        purl = apps['ProposalSnap']['app_url'] or APP_URLS['ProposalSnap']
        rp = fetch_api(purl, f'/api/proposals?company_name={urlquote(selected["name"])}', api_key)
        if rp: proposals = rp.get('proposals', [])
    # Fallback demo data for Bloom Studio
    if not proposals and selected['name'].lower().strip() == 'bloom studio':
        proposals = [
            {'title': 'Bloom Studio — Brand Identity Package', 'client_name': 'Varnam Artboutique', 'status': 'generated', 'created_at': '2026-02-10', 'num_slides': 12},
            {'title': 'Wedding Decor — Lotus Theme Collection', 'client_name': 'Priya & Arjun', 'status': 'generated', 'created_at': '2026-01-22', 'num_slides': 8},
            {'title': 'Corporate Art Workshop — Q1 Team Building', 'client_name': 'TechNova Solutions', 'status': 'generated', 'created_at': '2026-02-14', 'num_slides': 6},
        ]

    if is_demo:
        dash_urls = {k: v + '/demo' for k, v in APP_URLS.items()}
    else:
        sso_token = generate_sso_token(user['email'])
        dash_urls = {k: v + '/auto-login?token=' + sso_token for k, v in APP_URLS.items()}

    return render_template('dashboard.html', user=user, curr=curr, is_demo=is_demo,
        companies=companies, selected=selected, apps=apps, app_urls=dash_urls,
        invoices=invoices[:5], contracts=active_contracts[:5], proposals=proposals,
        total_invoiced=total_invoiced, total_paid=total_paid, total_unpaid=total_unpaid,
        total_overdue=total_overdue, total_expenses=total_expenses,
        total_payroll=total_payroll, total_payroll_gross=total_payroll_gross, total_payroll_net=total_payroll_net,
        contract_value=contract_value,
        revenue=revenue, costs=costs, profit=profit,
        expense_cats=expense_cats, monthly=monthly, mcv=mcv,
        cash_flow=cash_flow, balance_sheet=balance_sheet)

# ── Admin ───────────────────────────────────────────────────────
@app.route('/admin')
@login_required
def admin_dashboard():
    user = get_user()
    if not user['is_superadmin']:
        flash('Admin only', 'error'); return redirect(url_for('dashboard'))

    conn = get_db(); cur = conn.cursor()

    # Fetch all users and companies
    cur.execute('SELECT * FROM users ORDER BY created_at')
    all_users = cur.fetchall()

    cur.execute('SELECT * FROM companies ORDER BY name')
    companies = cur.fetchall()

    # Build company -> owner map via company_users join table
    cur.execute('''SELECT cu.company_id, cu.user_id, u.email, u.name as uname
                   FROM company_users cu JOIN users u ON cu.user_id=u.id
                   WHERE cu.role=%s''', ('owner',))
    ownership_rows = cur.fetchall()
    company_owner = {r['company_id']: r for r in ownership_rows}

    # Fallback: use owner_email field on company
    user_by_email = {u['email']: u for u in all_users}

    summaries = {}
    for c in companies:
        cur.execute('SELECT * FROM company_apps WHERE company_id=%s', (c['id'],))
        apps_list = cur.fetchall()
        apps = {a['app_name']: a for a in apps_list}
        api_key = c.get('owner_email', user['email'])
        s = {'apps': apps, 'invoiced':0, 'paid':0, 'expenses':0, 'payroll':0, 'contracts':0}

        for an, ai in apps.items():
            url = ai.get('app_url','') or APP_URLS.get(an,'')
            if an == 'InvoiceSnap':
                r = fetch_api(url, '/api/invoices', api_key)
                if r:
                    s['invoiced'] = sum(float(i.get('total',0) or 0) for i in r.get('invoices',[]))
                    s['paid'] = sum(float(i.get('total',0) or 0) for i in r.get('invoices',[]) if i.get('status')=='paid')
            elif an == 'ExpenseSnap':
                ecid = ''
                r2 = fetch_api(url, '/api/companies/external', api_key)
                if r2:
                    for ec in r2.get('companies',[]):
                        if ec['name'].lower().strip() == c['name'].lower().strip():
                            ecid = str(ec['id']); break
                ep = '/api/expenses/external'
                if ecid: ep += f'?company_id={ecid}'
                r = fetch_api(url, ep, api_key)
                if r: s['expenses'] = sum(float(e.get('total',0) or e.get('amount',0) or 0) for e in r.get('expenses',[]))
            elif an == 'ContractSnap':
                r = fetch_api(url, '/api/contracts', api_key)
                if r: s['contracts'] = sum(float(ct.get('total_value',0) or 0) for ct in r.get('contracts',[]))
            elif an == 'PayslipSnap':
                r = fetch_api(url, '/api/payroll', api_key)
                if r: s['payroll'] = sum(
                    float(p.get('gross_earnings',0) or 0) +
                    float(p.get('pf_employer',0) or 0) +
                    float(p.get('esi_employer',0) or 0)
                    for p in r.get('payslips',[]))
        s['profit'] = s['paid'] - s['expenses'] - s['payroll']
        summaries[c['id']] = s

    # Group companies by owner
    from collections import defaultdict
    users_companies = defaultdict(list)
    unassigned = []
    for c in companies:
        owner = company_owner.get(c['id'])
        if owner:
            users_companies[owner['user_id']].append(c)
        elif c.get('owner_email') and c['owner_email'] in user_by_email:
            u = user_by_email[c['owner_email']]
            users_companies[u['id']].append(c)
        else:
            unassigned.append(c)

    # Build user groups for template
    user_groups = []
    for u in all_users:
        ucos = users_companies.get(u['id'], [])
        if ucos:
            user_groups.append({'user': u, 'companies': ucos})
    if unassigned:
        user_groups.append({'user': {'name': 'Unassigned', 'email': '—', 'is_superadmin': False}, 'companies': unassigned})

    totals = {'companies': len(companies), 'users': len(all_users)}
    for k in ('invoiced','paid','expenses','payroll','contracts'):
        totals[k] = sum(s[k] for s in summaries.values())
    totals['profit'] = totals['paid'] - totals['expenses'] - totals['payroll']
    conn.close()

    return render_template('admin.html', user=user, user_groups=user_groups,
                         summaries=summaries, totals=totals, cs=cs, app_urls=APP_URLS)

@app.route('/settings', methods=['GET','POST'])
@login_required
def settings():
    user = get_user()
    if request.method == 'POST':
        conn = get_db(); cur = conn.cursor()
        cur.execute('UPDATE users SET name=%s, currency=%s WHERE id=%s',
                   (request.form.get('name',''), request.form.get('currency','INR'), user['id']))
        conn.close(); flash('Saved!', 'success'); return redirect(url_for('settings'))
    return render_template('settings.html', user=user, app_urls=APP_URLS)

@app.route('/settings/urls', methods=['POST'])
@login_required
def save_urls():
    user = get_user()
    if not user['is_superadmin']:
        flash('Admin only', 'error'); return redirect(url_for('settings'))
    conn = get_db(); cur = conn.cursor()
    for app_name in DEFAULT_URLS:
        url = request.form.get(app_name, '').strip()
        if url:
            cur.execute('''INSERT INTO app_settings (key, value) VALUES (%s, %s)
                          ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value''',
                       (f'url_{app_name}', url))
    conn.close()
    flash('App URLs saved!', 'success')
    return redirect(url_for('settings'))

@app.route('/drilldown/<app_name>')
@login_required
def drilldown(app_name):
    user = get_user()
    companies = get_user_companies(user)
    cid = request.args.get('company', '')
    selected = None
    for c in companies:
        if str(c['id']) == cid: selected = c; break
    if not selected and companies: selected = companies[0]
    if not selected: return redirect(url_for('apps_hub'))

    curr = cs(selected.get('currency', 'INR'))
    apps = get_company_apps(selected['id'])
    api_key = selected.get('owner_email', user['email'])
    data = []
    app_url = ''
    title = ''

    if app_name == 'invoices':
        title = 'Invoices'
        app_url = APP_URLS.get('InvoiceSnap', '')
        if 'InvoiceSnap' in apps:
            url = apps['InvoiceSnap']['app_url'] or APP_URLS['InvoiceSnap']
            r = fetch_api(url, f'/api/invoices?company_name={urlquote(selected["name"])}', api_key)
            if r: data = r.get('invoices', [])
            # Add formatted date
            for i in data:
                i['date'] = (i.get('issue_date') or '')[:10]

    elif app_name == 'contracts':
        title = 'Contracts'
        app_url = APP_URLS.get('ContractSnap', '')
        if 'ContractSnap' in apps:
            url = apps['ContractSnap']['app_url'] or APP_URLS['ContractSnap']
            r = fetch_api(url, f'/api/contracts?company_name={urlquote(selected["name"])}', api_key)
            if r: data = r.get('contracts', [])

    elif app_name == 'expenses':
        title = 'Expenses'
        app_url = APP_URLS.get('ExpenseSnap', '')
        if 'ExpenseSnap' in apps:
            url = apps['ExpenseSnap']['app_url'] or APP_URLS['ExpenseSnap']
            ecid = ''
            r2 = fetch_api(url, '/api/companies/external', api_key)
            if r2:
                for ec in r2.get('companies', []):
                    if ec['name'].lower().strip() == selected['name'].lower().strip():
                        ecid = str(ec['id']); break
            ep = '/api/expenses/external'
            if ecid: ep += f'?company_id={ecid}'
            r = fetch_api(url, ep, api_key)
            if r: data = r.get('expenses', [])

    elif app_name == 'payroll':
        title = 'Payroll'
        app_url = APP_URLS.get('PayslipSnap', '')
        if 'PayslipSnap' in apps:
            url = apps['PayslipSnap']['app_url'] or APP_URLS['PayslipSnap']
            r = fetch_api(url, f'/api/payroll?company_name={urlquote(selected["name"])}', api_key)
            if r: data = r.get('payslips', [])

    elif app_name == 'proposals':
        title = 'Proposals'
        app_url = APP_URLS.get('ProposalSnap', '')
        if 'ProposalSnap' in apps:
            url = apps['ProposalSnap']['app_url'] or APP_URLS['ProposalSnap']
            r = fetch_api(url, f'/api/proposals?company_name={urlquote(selected["name"])}', api_key)
            if r: data = r.get('proposals', [])
        if not data and selected['name'].lower().strip() == 'bloom studio':
            data = [
                {'title': 'Brand Identity Package', 'client_name': 'Varnam Artboutique', 'status': 'generated', 'created_at': '2026-02-10', 'num_slides': 12},
                {'title': 'Wedding Decor — Lotus Theme Collection', 'client_name': 'Priya & Arjun', 'status': 'generated', 'created_at': '2026-01-22', 'num_slides': 8},
                {'title': 'Corporate Art Workshop — Q1', 'client_name': 'TechNova Solutions', 'status': 'generated', 'created_at': '2026-02-14', 'num_slides': 6},
            ]
    return render_template('drilldown.html', app_name=app_name, title=title, company=selected,
                          data=data, curr=curr, app_url=app_url, user=user)

@app.route('/export/<app_name>')
@login_required
def export_excel(app_name):
    """Export drilldown data as Excel file."""
    user = get_user()
    companies = get_user_companies(user)
    cid = request.args.get('company', '')
    selected = None
    for c in companies:
        if str(c['id']) == cid: selected = c; break
    if not selected and companies: selected = companies[0]
    if not selected: return 'No company', 404

    curr = cs(selected.get('currency', 'INR'))
    apps = get_company_apps(selected['id'])
    api_key = selected.get('owner_email', user['email'])
    data = []

    # Fetch data (same logic as drilldown)
    if app_name == 'invoices' and 'InvoiceSnap' in apps:
        url = apps['InvoiceSnap']['app_url'] or APP_URLS['InvoiceSnap']
        r = fetch_api(url, f'/api/invoices?company_name={urlquote(selected["name"])}', api_key)
        if r: data = r.get('invoices', [])
    elif app_name == 'contracts' and 'ContractSnap' in apps:
        url = apps['ContractSnap']['app_url'] or APP_URLS['ContractSnap']
        r = fetch_api(url, f'/api/contracts?company_name={urlquote(selected["name"])}', api_key)
        if r: data = r.get('contracts', [])
    elif app_name == 'expenses' and 'ExpenseSnap' in apps:
        url = apps['ExpenseSnap']['app_url'] or APP_URLS['ExpenseSnap']
        ecid = ''
        r2 = fetch_api(url, '/api/companies/external', api_key)
        if r2:
            for ec in r2.get('companies', []):
                if ec['name'].lower().strip() == selected['name'].lower().strip():
                    ecid = str(ec['id']); break
        ep = '/api/expenses/external'
        if ecid: ep += f'?company_id={ecid}'
        r = fetch_api(url, ep, api_key)
        if r: data = r.get('expenses', [])
    elif app_name == 'payroll' and 'PayslipSnap' in apps:
        url = apps['PayslipSnap']['app_url'] or APP_URLS['PayslipSnap']
        r = fetch_api(url, f'/api/payroll?company_name={urlquote(selected["name"])}', api_key)
        if r: data = r.get('payslips', [])

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = app_name.capitalize()

    # Styles
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14)
    money_fmt = '#,##0.00'
    thin_border = Border(
        bottom=Side(style='thin', color='DDDDDD')
    )

    # Title row
    ws.merge_cells('A1:E1')
    ws['A1'].value = f'{selected["name"]} — {app_name.capitalize()}'
    ws['A1'].font = title_font
    ws.append([])  # blank row

    if app_name == 'invoices':
        headers = ['Invoice #', 'Client', 'Date', 'Status', f'Total ({curr.strip()})']
        ws.append(headers)
        for col in range(1, 6):
            cell = ws.cell(row=3, column=col)
            cell.font = header_font; cell.fill = header_fill
        for i in data:
            ws.append([i.get('invoice_number',''), i.get('client_name',''),
                      str(i.get('issue_date',''))[:10], (i.get('status','') or '').capitalize(),
                      float(i.get('total',0) or 0)])
            ws.cell(row=ws.max_row, column=5).number_format = money_fmt
        # Summary
        ws.append([])
        paid = sum(float(i.get('total',0) or 0) for i in data if i.get('status')=='paid')
        total = sum(float(i.get('total',0) or 0) for i in data)
        ws.append(['', '', '', 'Total Invoiced:', total])
        ws.append(['', '', '', 'Total Paid:', paid])
        ws.append(['', '', '', 'Outstanding:', total - paid])
        for r in range(ws.max_row-2, ws.max_row+1):
            ws.cell(row=r, column=4).font = Font(bold=True)
            ws.cell(row=r, column=5).number_format = money_fmt
            ws.cell(row=r, column=5).font = Font(bold=True)
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 18

    elif app_name == 'contracts':
        headers = ['Contract #', 'Title', 'Client', 'Status', f'Value ({curr.strip()})']
        ws.append(headers)
        for col in range(1, 6):
            cell = ws.cell(row=3, column=col)
            cell.font = header_font; cell.fill = header_fill
        for c in data:
            ws.append([c.get('contract_number',''), c.get('title',''), c.get('client_name',''),
                      (c.get('status','') or '').capitalize(), float(c.get('total_value',0) or 0)])
            ws.cell(row=ws.max_row, column=5).number_format = money_fmt
        ws.append([])
        tv = sum(float(c.get('total_value',0) or 0) for c in data)
        ws.append(['', '', '', 'Total Value:', tv])
        ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=5).number_format = money_fmt
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 18

    elif app_name == 'expenses':
        headers = ['Date', 'Vendor', 'Category', 'Currency', 'Total']
        ws.append(headers)
        for col in range(1, 6):
            cell = ws.cell(row=3, column=col)
            cell.font = header_font; cell.fill = header_fill
        for e in data:
            ws.append([e.get('date', e.get('receipt_date','')), e.get('vendor', e.get('merchant','')),
                      e.get('category',''), e.get('currency',''), float(e.get('total', e.get('amount',0)) or 0)])
            ws.cell(row=ws.max_row, column=5).number_format = money_fmt
        ws.append([])
        te = sum(float(e.get('total', e.get('amount',0)) or 0) for e in data)
        ws.append(['', '', '', 'Total:', te])
        ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=5).number_format = money_fmt
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 16

    elif app_name == 'payroll':
        headers = ['Employee', 'Period', f'Gross ({curr.strip()})', f'Deductions ({curr.strip()})', f'Net Pay ({curr.strip()})']
        ws.append(headers)
        for col in range(1, 6):
            cell = ws.cell(row=3, column=col)
            cell.font = header_font; cell.fill = header_fill
        for p in data:
            ws.append([p.get('emp_name', p.get('employee_name','')),
                      f"{p.get('month','')}/{p.get('year','')}",
                      float(p.get('gross_earnings',0) or 0),
                      float(p.get('total_deductions',0) or 0),
                      float(p.get('net_pay',0) or 0)])
            for c in [3,4,5]: ws.cell(row=ws.max_row, column=c).number_format = money_fmt
        ws.append([])
        tg = sum(float(p.get('gross_earnings',0) or 0) for p in data)
        td = sum(float(p.get('total_deductions',0) or 0) for p in data)
        tn = sum(float(p.get('net_pay',0) or 0) for p in data)
        ws.append(['', 'Totals:', tg, td, tn])
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
        for c in [3,4,5]:
            ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=c).number_format = money_fmt
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16

    # Apply borders to data rows
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=5):
        for cell in row:
            cell.border = thin_border

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f'{selected["name"]}_{app_name}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/all')
@login_required
def export_all():
    """Export all data for a company into one multi-sheet Excel file."""
    user = get_user()
    companies = get_user_companies(user)
    cid = request.args.get('company', '')
    selected = None
    for c in companies:
        if str(c['id']) == cid: selected = c; break
    if not selected and companies: selected = companies[0]
    if not selected: return 'No company', 404

    curr = cs(selected.get('currency', 'INR'))
    apps = get_company_apps(selected['id'])
    api_key = selected.get('owner_email', user['email'])

    wb = Workbook()
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    money_fmt = '#,##0.00'
    thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))

    first_sheet = True

    # Invoices
    if 'InvoiceSnap' in apps:
        url = apps['InvoiceSnap']['app_url'] or APP_URLS['InvoiceSnap']
        r = fetch_api(url, f'/api/invoices?company_name={urlquote(selected["name"])}', api_key)
        invoices = r.get('invoices', []) if r else []
        if first_sheet: ws = wb.active; ws.title = 'Invoices'; first_sheet = False
        else: ws = wb.create_sheet('Invoices')
        ws.append(['Invoice #', 'Client', 'Date', 'Status', f'Total ({curr.strip()})'])
        for col in range(1, 6):
            ws.cell(row=1, column=col).font = header_font; ws.cell(row=1, column=col).fill = header_fill
        for i in invoices:
            ws.append([i.get('invoice_number',''), i.get('client_name',''), str(i.get('issue_date',''))[:10],
                      (i.get('status','') or '').capitalize(), float(i.get('total',0) or 0)])
            ws.cell(row=ws.max_row, column=5).number_format = money_fmt
        ws.column_dimensions['A'].width = 16; ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 14; ws.column_dimensions['D'].width = 12; ws.column_dimensions['E'].width = 16

    # Contracts
    if 'ContractSnap' in apps:
        url = apps['ContractSnap']['app_url'] or APP_URLS['ContractSnap']
        r = fetch_api(url, f'/api/contracts?company_name={urlquote(selected["name"])}', api_key)
        contracts = r.get('contracts', []) if r else []
        if first_sheet: ws = wb.active; ws.title = 'Contracts'; first_sheet = False
        else: ws = wb.create_sheet('Contracts')
        ws.append(['Contract #', 'Title', 'Client', 'Status', f'Value ({curr.strip()})'])
        for col in range(1, 6):
            ws.cell(row=1, column=col).font = header_font; ws.cell(row=1, column=col).fill = header_fill
        for c in contracts:
            ws.append([c.get('contract_number',''), c.get('title',''), c.get('client_name',''),
                      (c.get('status','') or '').capitalize(), float(c.get('total_value',0) or 0)])
            ws.cell(row=ws.max_row, column=5).number_format = money_fmt
        ws.column_dimensions['A'].width = 16; ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 22; ws.column_dimensions['D'].width = 12; ws.column_dimensions['E'].width = 16

    # Expenses
    if 'ExpenseSnap' in apps:
        url = apps['ExpenseSnap']['app_url'] or APP_URLS['ExpenseSnap']
        ecid = ''
        r2 = fetch_api(url, '/api/companies/external', api_key)
        if r2:
            for ec in r2.get('companies', []):
                if ec['name'].lower().strip() == selected['name'].lower().strip():
                    ecid = str(ec['id']); break
        ep = '/api/expenses/external'
        if ecid: ep += f'?company_id={ecid}'
        r = fetch_api(url, ep, api_key)
        expenses = r.get('expenses', []) if r else []
        if first_sheet: ws = wb.active; ws.title = 'Expenses'; first_sheet = False
        else: ws = wb.create_sheet('Expenses')
        ws.append(['Date', 'Vendor', 'Category', 'Currency', 'Total'])
        for col in range(1, 6):
            ws.cell(row=1, column=col).font = header_font; ws.cell(row=1, column=col).fill = header_fill
        for e in expenses:
            ws.append([e.get('date', e.get('receipt_date','')), e.get('vendor', e.get('merchant','')),
                      e.get('category',''), e.get('currency',''), float(e.get('total', e.get('amount',0)) or 0)])
            ws.cell(row=ws.max_row, column=5).number_format = money_fmt
        ws.column_dimensions['A'].width = 14; ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 20; ws.column_dimensions['D'].width = 10; ws.column_dimensions['E'].width = 16

    # Payroll
    if 'PayslipSnap' in apps:
        url = apps['PayslipSnap']['app_url'] or APP_URLS['PayslipSnap']
        r = fetch_api(url, f'/api/payroll?company_name={urlquote(selected["name"])}', api_key)
        payslips = r.get('payslips', []) if r else []
        if first_sheet: ws = wb.active; ws.title = 'Payroll'; first_sheet = False
        else: ws = wb.create_sheet('Payroll')
        ws.append(['Employee', 'Period', f'Gross ({curr.strip()})', f'Deductions ({curr.strip()})', f'Net Pay ({curr.strip()})'])
        for col in range(1, 6):
            ws.cell(row=1, column=col).font = header_font; ws.cell(row=1, column=col).fill = header_fill
        for p in payslips:
            ws.append([p.get('emp_name', p.get('employee_name','')), f"{p.get('month','')}/{p.get('year','')}",
                      float(p.get('gross_earnings',0) or 0), float(p.get('total_deductions',0) or 0),
                      float(p.get('net_pay',0) or 0)])
            for c in [3,4,5]: ws.cell(row=ws.max_row, column=c).number_format = money_fmt
        ws.column_dimensions['A'].width = 24; ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 16; ws.column_dimensions['D'].width = 16; ws.column_dimensions['E'].width = 16

    if first_sheet:
        ws = wb.active; ws.title = 'No Data'; ws.append(['No data found for this company'])

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f'{selected["name"]}_VarnamSuite_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/seed-test-data', methods=['POST'])
@login_required
def seed_test_data():
    """Create test company 'Bloom Studio' with data across all apps."""
    user = get_user()
    if not user['is_superadmin']:
        flash('Admin only', 'error'); return redirect(url_for('admin_dashboard'))

    api_key = user['email']
    results = {}

    for app_name in ['ExpenseSnap', 'InvoiceSnap', 'ContractSnap', 'PayslipSnap']:
        url = APP_URLS.get(app_name, '')
        if not url: results[app_name] = 'No URL'; continue
        try:
            r = requests.post(url.rstrip('/') + '/api/seed-test-data',
                            headers={'X-API-Key': api_key}, timeout=30)
            results[app_name] = r.json() if r.status_code == 200 else f'Error {r.status_code}: {r.text[:100]}'
        except Exception as e:
            results[app_name] = f'Failed: {str(e)[:100]}'

    # Now register Bloom Studio in Varnam Suite hub
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT * FROM companies WHERE LOWER(name)=LOWER(%s)', ('Bloom Studio',))
    if not cur.fetchone():
        cur.execute('INSERT INTO companies (name,currency,owner_email) VALUES (%s,%s,%s) RETURNING *',
                   ('Bloom Studio', 'INR', user['email']))
        company = cur.fetchone()
        for app_name in ['ExpenseSnap', 'InvoiceSnap', 'ContractSnap', 'PayslipSnap']:
            url = APP_URLS.get(app_name, '')
            cur.execute('''INSERT INTO company_apps (company_id,app_name,app_company_name,app_url)
                          VALUES (%s,%s,%s,%s) ON CONFLICT (company_id,app_name) DO NOTHING''',
                       (company['id'], app_name, 'Bloom Studio', url))
        cur.execute('INSERT INTO company_users (company_id,user_id,role) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING',
                   (company['id'], user['id'], 'owner'))
    conn.close()

    flash(f'Test data created! Results: {results}', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/sync', methods=['POST'])
@login_required
def sync_all():
    """Pull existing companies from all Varnam Suite apps."""
    user = get_user()
    if not user['is_superadmin']:
        flash('Admin only', 'error'); return redirect(url_for('dashboard'))

    api_key = user['email']
    added = 0; errors = []

    def register_company(app_name, company_name, currency, app_url):
        """Direct DB insert instead of self-HTTP call."""
        nonlocal added
        if not company_name: return
        conn = get_db(); cur = conn.cursor()
        cur.execute('SELECT COUNT(*) as cnt FROM companies')
        if cur.fetchone()['cnt'] >= MAX_COMPANIES:
            conn.close(); return
        cur.execute('SELECT * FROM companies WHERE LOWER(name)=LOWER(%s)', (company_name,))
        company = cur.fetchone()
        if not company:
            cur.execute('INSERT INTO companies (name,currency,owner_email) VALUES (%s,%s,%s) RETURNING *',
                       (company_name, currency, user['email']))
            company = cur.fetchone()
            added += 1
        cur.execute('''INSERT INTO company_apps (company_id,app_name,app_company_name,app_url)
                      VALUES (%s,%s,%s,%s) ON CONFLICT (company_id,app_name)
                      DO UPDATE SET app_company_name=EXCLUDED.app_company_name, app_url=EXCLUDED.app_url''',
                   (company['id'], app_name, company_name, app_url))
        cur.execute('''INSERT INTO company_users (company_id,user_id,role)
                      VALUES (%s,%s,'owner') ON CONFLICT DO NOTHING''', (company['id'], user['id']))
        conn.close()

    # Sync from ExpenseSnap
    r = fetch_api(APP_URLS['ExpenseSnap'], '/api/companies/external', api_key)
    if r:
        for ec in r.get('companies', []):
            register_company('ExpenseSnap', ec['name'], ec.get('home_currency', 'INR'), APP_URLS['ExpenseSnap'])
    else:
        errors.append('ExpenseSnap: could not reach API')

    # Sync from InvoiceSnap
    r = fetch_api(APP_URLS['InvoiceSnap'], '/api/invoices', api_key)
    if r:
        seen = set()
        for inv in r.get('invoices', []):
            cn = inv.get('company_name', '').strip()
            if cn and cn not in seen:
                seen.add(cn)
                register_company('InvoiceSnap', cn, user.get('currency', 'INR'), APP_URLS['InvoiceSnap'])
    else:
        errors.append('InvoiceSnap: could not reach API')

    # Sync from ContractSnap
    r = fetch_api(APP_URLS['ContractSnap'], '/api/contracts', api_key)
    if r:
        seen = set()
        for ct in r.get('contracts', []):
            cn = ct.get('company_name', '').strip()
            if cn and cn not in seen:
                seen.add(cn)
                register_company('ContractSnap', cn, user.get('currency', 'INR'), APP_URLS['ContractSnap'])
    else:
        errors.append('ContractSnap: could not reach API')

    # Sync from PayslipSnap
    r = fetch_api(APP_URLS['PayslipSnap'], '/api/payroll', api_key)
    if r:
        cn = user.get('name', '').strip() or 'Default'
        # PayslipSnap doesn't have company_name per payslip, skip for now
    else:
        errors.append('PayslipSnap: could not reach API')

    msg = f'Synced! {added} new companies added.'
    if errors:
        msg += f' Issues: {"; ".join(errors)}'
        flash(msg, 'error' if added == 0 else 'success')
    else:
        flash(msg, 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/diagnose')
@login_required
def diagnose():
    user = get_user()
    results = {}
    api_key = user['email']
    endpoints = {
        'ExpenseSnap': ['/api/companies/external', '/api/expenses/external'],
        'InvoiceSnap': ['/api/invoices'],
        'ContractSnap': ['/api/contracts'],
        'PayslipSnap': ['/api/payroll'],
    }
    for an, url in APP_URLS.items():
        eps = endpoints.get(an, [])
        if not eps:
            results[an] = {'url': url, 'status': 'N/A', 'detail': 'No API endpoint'}
            continue
        for ep in eps:
            try:
                r = requests.get(url.rstrip('/') + ep,
                    headers={'X-API-Key': api_key}, timeout=30)
                data = r.text[:300]
                results[f'{an} {ep}'] = {'url': url + ep, 'status': r.status_code, 'detail': data}
            except Exception as e:
                results[f'{an} {ep}'] = {'url': url + ep, 'status': 'Error', 'detail': str(e)[:300]}
    results['_info'] = {'url': '', 'status': 'Info',
        'detail': f'API Key (email): {api_key} | Is superadmin: {user.get("is_superadmin")}'}
    return render_template('diagnose.html', user=user, results=results)


@app.route('/health')
def health():
    return __import__('flask').jsonify({'status': 'ok', 'app': 'financesnap'})

# ── Bank Reconciliation ─────────────────────────────────────────

@app.route('/reconcile/sample-csv')
@login_required
def reconcile_sample_csv():
    """Download a sample bank statement CSV for testing reconciliation."""
    import io
    sample = """Date,Description,Amount,Balance
2026-01-03,Opening Balance,,50000.00
2026-01-05,CLIENT PAYMENT - TechNova Solutions,100300.00,150300.00
2026-01-07,RENT - Office Space Jan 2026,-25000.00,125300.00
2026-01-08,AWS Cloud Services,-4200.00,121100.00
2026-01-10,CLIENT PAYMENT - Priya Wellness Spa,41300.00,162400.00
2026-01-12,SWIGGY OFFICE LUNCH,-1850.00,160550.00
2026-01-14,ADOBE CREATIVE CLOUD,-3540.00,157010.00
2026-01-15,SALARY TRANSFER - Rahul Sharma,-45000.00,112010.00
2026-01-16,SALARY TRANSFER - Meena Iyer,-38000.00,74010.00
2026-01-18,CLIENT PAYMENT - Green Earth Organics,33040.00,107050.00
2026-01-20,GOOGLE WORKSPACE,-1200.00,105850.00
2026-01-21,OLA BUSINESS TRAVEL,-2300.00,103550.00
2026-01-22,STATIONERY - OFFICE DEPOT,-890.00,102660.00
2026-01-24,ZOOM SUBSCRIPTION,-1499.00,101161.00
2026-01-25,CLIENT PAYMENT - Meridian Architects,88500.00,189661.00
2026-01-27,FACEBOOK ADS,-8500.00,181161.00
2026-01-28,ELECTRICITY BILL,-3200.00,177961.00
2026-01-29,INTERNET - JIO FIBER,-1499.00,176462.00
2026-01-30,PERSONAL TRANSFER TO SAVINGS,-20000.00,156462.00
2026-01-31,BANK CHARGES,-250.00,156212.00
"""
    buf = io.BytesIO(sample.encode('utf-8'))
    buf.seek(0)
    from flask import send_file
    return send_file(buf, mimetype='text/csv',
                     as_attachment=True,
                     download_name='varnam_sample_bank_statement.csv')

@app.route('/reconcile')
@login_required
def reconcile():
    user = get_user()
    companies = get_user_companies(user)
    cid = request.args.get('company', '')
    selected = next((c for c in companies if str(c['id']) == cid), companies[0] if companies else None)
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT * FROM bank_statements WHERE user_id=%s ORDER BY uploaded_at DESC LIMIT 20', (user['id'],))
    statements = cur.fetchall()
    conn.close()
    return render_template('reconcile.html', user=user, companies=companies,
                           selected=selected, statements=statements)

@app.route('/reconcile/parse', methods=['POST'])
@login_required
def reconcile_parse():
    """Step 1: Upload CSV, use AI to detect columns, return preview."""
    user = get_user()
    if 'csv_file' not in request.files or not request.files['csv_file'].filename:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['csv_file']
    if not f.filename.lower().endswith('.csv'):
        return jsonify({'error': 'Please upload a CSV file'}), 400

    content = f.read().decode('utf-8', errors='replace')
    lines = content.strip().split('\n')
    if len(lines) < 2:
        return jsonify({'error': 'CSV appears empty'}), 400

    # Send first 10 rows to Claude for column detection
    sample = '\n'.join(lines[:min(10, len(lines))])
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'))
        msg = client.messages.create(
            model='claude-opus-4-5',
            max_tokens=800,
            system='''You are a bank statement CSV parser. Analyze the CSV sample and identify columns.
Return ONLY valid JSON with this structure:
{
  "date_col": <column index 0-based or null>,
  "desc_col": <column index 0-based or null>,
  "amount_col": <column index 0-based or null>,
  "debit_col": <column index if debit is separate column, else null>,
  "credit_col": <column index if credit is separate column, else null>,
  "has_header": <true/false>,
  "date_format": "<detected format like DD/MM/YYYY or YYYY-MM-DD>",
  "confidence": "<high/medium/low>",
  "notes": "<brief explanation>"
}
If amount is a single signed column, set amount_col and leave debit_col/credit_col null.
If debits and credits are separate columns, set debit_col and credit_col, leave amount_col null.''',
            messages=[{'role': 'user', 'content': f'Parse this bank statement CSV:\n{sample}'}]
        )
        import json
        mapping = json.loads(msg.content[0].text)
    except Exception as e:
        mapping = {'date_col': 0, 'desc_col': 1, 'amount_col': 2, 'debit_col': None,
                   'credit_col': None, 'has_header': True, 'date_format': 'auto',
                   'confidence': 'low', 'notes': f'AI detection failed: {str(e)}'}

    # Parse rows using detected mapping
    import csv, io
    reader = csv.reader(io.StringIO(content))
    rows = list(reader)
    header = rows[0] if mapping.get('has_header') else []
    data_rows = rows[1:] if mapping.get('has_header') else rows

    # Store raw CSV in session for step 2
    session['recon_csv'] = content
    session['recon_mapping'] = mapping

    preview = []
    for row in data_rows[:20]:
        if not any(row): continue
        try:
            date_val = row[mapping['date_col']].strip() if mapping.get('date_col') is not None and mapping['date_col'] < len(row) else ''
            desc_val = row[mapping['desc_col']].strip() if mapping.get('desc_col') is not None and mapping['desc_col'] < len(row) else ''
            if mapping.get('debit_col') is not None and mapping.get('credit_col') is not None:
                debit = float(row[mapping['debit_col']].replace(',','').strip() or 0) if mapping['debit_col'] < len(row) else 0
                credit = float(row[mapping['credit_col']].replace(',','').strip() or 0) if mapping['credit_col'] < len(row) else 0
                amount = credit - debit if credit > 0 else -debit
            else:
                raw = row[mapping['amount_col']].replace(',','').strip() if mapping.get('amount_col') is not None and mapping['amount_col'] < len(row) else '0'
                amount = float(raw or 0)
            preview.append({'date': date_val, 'description': desc_val, 'amount': round(amount, 2)})
        except:
            continue

    return jsonify({'mapping': mapping, 'header': header, 'preview': preview,
                    'total_rows': len(data_rows), 'columns': header or [f'Col {i}' for i in range(len(rows[0]) if rows else 0)]})

@app.route('/reconcile/match', methods=['POST'])
@login_required
def reconcile_match():
    """Step 2: With confirmed mapping, parse all rows and match against invoices/expenses/payroll."""
    user = get_user()
    data = request.get_json() or {}
    mapping = data.get('mapping') or session.get('recon_mapping', {})
    company_name = data.get('company_name', '')
    currency = data.get('currency', 'USD')
    content = session.get('recon_csv', '')
    if not content:
        return jsonify({'error': 'No CSV in session. Please re-upload.'}), 400

    import csv, io
    reader = csv.reader(io.StringIO(content))
    rows = list(reader)
    data_rows = rows[1:] if mapping.get('has_header') else rows

    # Parse all rows
    transactions = []
    for row in data_rows:
        if not any(row): continue
        try:
            date_val = row[mapping['date_col']].strip() if mapping.get('date_col') is not None and mapping['date_col'] < len(row) else ''
            desc_val = row[mapping['desc_col']].strip() if mapping.get('desc_col') is not None and mapping['desc_col'] < len(row) else ''
            if mapping.get('debit_col') is not None and mapping.get('credit_col') is not None:
                debit = float(row[mapping['debit_col']].replace(',','').strip() or 0) if mapping['debit_col'] < len(row) else 0
                credit = float(row[mapping['credit_col']].replace(',','').strip() or 0) if mapping['credit_col'] < len(row) else 0
                amount = credit - debit if credit > 0 else -debit
            else:
                raw = row[mapping['amount_col']].replace(',','').strip() if mapping.get('amount_col') is not None and mapping['amount_col'] < len(row) else '0'
                amount = float(raw or 0)
            txn_type = 'credit' if amount > 0 else 'debit'
            transactions.append({'date': date_val[:10], 'description': desc_val,
                                  'amount': abs(round(amount, 2)), 'type': txn_type, 'status': 'unmatched',
                                  'matched_type': '', 'matched_id': '', 'matched_label': ''})
        except:
            continue

    # Fetch source data for matching
    api_key = user['email']
    apps_data = get_user_apps(user)

    # Pull invoices
    invoices = []
    if 'InvoiceSnap' in apps_data:
        url = apps_data['InvoiceSnap']['app_url'] or APP_URLS['InvoiceSnap']
        r = fetch_api(url, f'/api/invoices?company_name={urlquote(company_name)}', api_key)
        if r: invoices = [i for i in r.get('invoices', []) if str(i.get('status','')).lower() == 'paid']

    # Pull expenses
    expenses = []
    if 'ExpenseSnap' in apps_data:
        url = apps_data['ExpenseSnap']['app_url'] or APP_URLS['ExpenseSnap']
        r = fetch_api(url, f'/api/expenses/external?company_name={urlquote(company_name)}', api_key)
        if r: expenses = r.get('expenses', [])

    # Pull payroll
    payroll = []
    if 'PayslipSnap' in apps_data:
        url = apps_data['PayslipSnap']['app_url'] or APP_URLS['PayslipSnap']
        r = fetch_api(url, f'/api/payroll?company_name={urlquote(company_name)}', api_key)
        if r: payroll = r.get('payslips', [])

    # Match each transaction
    from datetime import datetime, timedelta
    def parse_date(s):
        for fmt in ('%Y-%m-%d','%d/%m/%Y','%m/%d/%Y','%d-%m-%Y','%Y%m%d'):
            try: return datetime.strptime(s[:10], fmt)
            except: continue
        return None

    def dates_close(d1, d2, days=3):
        try: return abs((parse_date(d1) - parse_date(d2)).days) <= days
        except: return False

    def amounts_match(a1, a2, pct=0.01):
        return abs(float(a1) - float(a2)) <= max(0.02, float(a2) * pct)

    for txn in transactions:
        if txn['type'] == 'credit':
            # Match against paid invoices
            for inv in invoices:
                inv_amount = float(inv.get('total_amount') or inv.get('amount') or 0)
                inv_date = str(inv.get('paid_at') or inv.get('issue_date') or '')[:10]
                if amounts_match(txn['amount'], inv_amount) and dates_close(txn['date'], inv_date):
                    txn['status'] = 'matched'
                    txn['matched_type'] = 'invoice'
                    txn['matched_id'] = str(inv.get('id',''))
                    txn['matched_label'] = f"Invoice #{inv.get('invoice_number','?')} — {inv.get('client_name','')}"
                    break
        else:
            # Match against expenses first
            for exp in expenses:
                exp_amount = float(exp.get('total') or exp.get('amount') or 0)
                exp_date = str(exp.get('date') or exp.get('receipt_date') or '')[:10]
                if amounts_match(txn['amount'], exp_amount) and dates_close(txn['date'], exp_date):
                    txn['status'] = 'matched'
                    txn['matched_type'] = 'expense'
                    txn['matched_id'] = str(exp.get('id',''))
                    txn['matched_label'] = f"{exp.get('vendor','Expense')} — {exp.get('category','')}"
                    break
            # If still unmatched, try payroll
            if txn['status'] == 'unmatched':
                for slip in payroll:
                    net = float(slip.get('net_salary') or slip.get('net_pay') or 0)
                    slip_month = str(slip.get('month',''))
                    slip_year = str(slip.get('year',''))
                    txn_month = txn['date'][5:7].lstrip('0') if len(txn['date']) >= 7 else ''
                    txn_year = txn['date'][:4] if len(txn['date']) >= 4 else ''
                    if amounts_match(txn['amount'], net) and slip_month == txn_month and slip_year == txn_year:
                        txn['status'] = 'matched'
                        txn['matched_type'] = 'payroll'
                        txn['matched_id'] = str(slip.get('id',''))
                        txn['matched_label'] = f"Payroll — {slip.get('emp_name','Employee')} {slip_month}/{slip_year}"
                        break

    matched = sum(1 for t in transactions if t['status'] == 'matched')
    session['recon_transactions'] = transactions
    session['recon_company'] = company_name
    session['recon_currency'] = currency
    return jsonify({'transactions': transactions, 'total': len(transactions),
                    'matched': matched, 'unmatched': len(transactions) - matched})

@app.route('/reconcile/confirm', methods=['POST'])
@login_required
def reconcile_confirm():
    """Step 3: Save statement + push new expenses to ExpenseSnap."""
    user = get_user()
    data = request.get_json() or {}
    transactions = data.get('transactions', session.get('recon_transactions', []))
    company_name = data.get('company_name', session.get('recon_company', ''))
    currency = data.get('currency', session.get('recon_currency', 'USD'))
    filename = data.get('filename', 'bank_statement.csv')
    new_expenses = [t for t in transactions if t.get('status') == 'new_expense']
    ignored = [t for t in transactions if t.get('status') == 'ignored']
    matched = [t for t in transactions if t.get('status') == 'matched']

    # Push new expenses to ExpenseSnap
    pushed = []
    apps_data = get_user_apps(user)
    if new_expenses and 'ExpenseSnap' in apps_data:
        url = apps_data['ExpenseSnap']['app_url'] or APP_URLS['ExpenseSnap']
        for txn in new_expenses:
            try:
                import urllib.request, json as jsonlib
                payload = jsonlib.dumps({
                    'date': txn['date'], 'amount': txn['amount'],
                    'description': txn.get('description','Bank Transaction'),
                    'category': txn.get('category','Other'),
                    'company_name': company_name, 'currency': currency,
                    'source': 'bank_recon'
                }).encode()
                req = urllib.request.Request(f"{url}/api/expenses/create-external",
                    data=payload, headers={'Content-Type':'application/json','X-API-Key': user['email']},
                    method='POST')
                with urllib.request.urlopen(req, timeout=10) as resp:
                    result = jsonlib.loads(resp.read())
                    txn['expense_snap_id'] = result.get('expense_id','')
                    pushed.append(txn)
            except Exception as e:
                txn['push_error'] = str(e)

    # Save statement record
    conn = get_db(); cur = conn.cursor()
    cur.execute('''INSERT INTO bank_statements (user_id, company_name, filename, row_count, matched_count, currency)
                   VALUES (%s,%s,%s,%s,%s,%s) RETURNING id''',
                (user['id'], company_name, filename, len(transactions), len(matched), currency))
    stmt_id = cur.fetchone()['id']

    for txn in transactions:
        cur.execute('''INSERT INTO bank_transactions
            (statement_id, txn_date, description, amount, txn_type, status, matched_type, matched_id, expense_snap_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)''',
            (stmt_id, txn['date'], txn['description'], txn['amount'], txn['type'],
             txn['status'], txn.get('matched_type',''), txn.get('matched_id',''),
             txn.get('expense_snap_id','')))
    conn.close()

    # Clear session
    session.pop('recon_csv', None)
    session.pop('recon_transactions', None)
    session.pop('recon_mapping', None)

    return jsonify({'success': True, 'statement_id': stmt_id,
                    'matched': len(matched), 'new_expenses_pushed': len(pushed),
                    'ignored': len(ignored)})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
